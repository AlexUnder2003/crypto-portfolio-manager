import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
import requests
import pandas as pd
import json
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Словарь для сопоставления идентификаторов
crypto_ids = {
    "bitcoin": "bitcoin",
    "ethereum": "ethereum",
    "litecoin": "litecoin",
    "render-token": "render-token",
    "ordinals": "ordinals",
    "toncoin": "the-open-network",
    "ethereum-classic": "ethereum-classic",
    "fetch-ai": "fetch-ai",
    "solana": "solana",
}

# Список криптовалют
cryptos = list(crypto_ids.keys())

# Проверка наличия файла с данными
data_file = "crypto_portfolio.json"
if os.path.exists(data_file):
    with open(data_file, "r") as f:
        portfolio = json.load(f)
else:
    portfolio = {crypto: 0 for crypto in cryptos}

# Кэш для хранения результатов запросов к API
price_cache = {}


# Функция для получения курса криптовалюты с обработкой ошибок и кэшированием
def get_crypto_price(crypto_id):
    try:
        # Проверяем кэш на наличие данных
        if crypto_id in price_cache:
            cached_data, expiry_time = price_cache[crypto_id]
            if datetime.now() < expiry_time:
                return cached_data
            else:
                del price_cache[crypto_id]

        url = f"https://api.coingecko.com/api/v3/simple/price?ids={crypto_id}&vs_currencies=rub"
        response = requests.get(url)

        # Проверяем статус ответа
        if response.status_code == 429:
            raise Exception(f"Too Many Requests: {response.text}")

        response.raise_for_status()  # Генерирует исключение для других ошибок HTTP

        data = response.json()
        if crypto_id in data:
            # Кэшируем результат на 1 минуту
            price_cache[crypto_id] = (
                data[crypto_id]["rub"],
                datetime.now() + timedelta(minutes=1),
            )
            return data[crypto_id]["rub"]
        else:
            print(f"Ошибка: Не удалось получить курс для {crypto_id}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при запросе курса для {crypto_id}: {e}")
        return None
    except Exception as e:
        print(f"Неожиданная ошибка при запросе курса для {crypto_id}: {e}")
        return None


# Функция для обновления портфеля
def update_portfolio():
    for crypto in cryptos:
        amount = entry_vars[crypto].get()
        try:
            portfolio[crypto] = float(amount)
        except ValueError:
            messagebox.showerror(
                "Ошибка", f"Введите корректное количество для {crypto}"
            )

    with open(data_file, "w") as f:
        json.dump(portfolio, f)

    messagebox.showinfo("Успех", "Портфель обновлен!")


# Функция для обновления списка листов в выпадающем списке
def update_sheet_list():
    excel_file = "crypto_portfolio.xlsx"
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        sheets = wb.sheetnames
    else:
        sheets = []
    sheet_combobox["values"] = sheets


# Функция для создания нового листа
def create_new_sheet():
    new_sheet = simpledialog.askstring(
        "Создание листа", "Введите имя нового листа:"
    )
    if new_sheet:
        excel_file = "crypto_portfolio.xlsx"
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            if new_sheet not in wb.sheetnames:
                wb.create_sheet(new_sheet)
                wb.save(excel_file)
                update_sheet_list()
                sheet_combobox.set(new_sheet)
                messagebox.showinfo(
                    "Успех", f"Лист '{new_sheet}' создан и выбран."
                )
            else:
                messagebox.showerror(
                    "Ошибка", f"Лист '{new_sheet}' уже существует."
                )
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = new_sheet
            wb.save(excel_file)
            update_sheet_list()
            sheet_combobox.set(new_sheet)
            messagebox.showinfo(
                "Успех", f"Лист '{new_sheet}' создан и выбран."
            )


# Функция для записи данных в Excel
def record_data():
    total_value_rub = 0
    crypto_values = {"Date": datetime.now().strftime("%Y-%m-%d")}
    all_prices_fetched = True

    for crypto in cryptos:
        amount = portfolio[crypto]
        if amount == 0:
            # Если количество равно 0, пропускаем получение курса и не добавляем значение
            crypto_values[crypto] = 0
            continue

        price = get_crypto_price(crypto_ids[crypto])
        if price is not None:
            value = amount * price
            total_value_rub += value
            crypto_values[crypto] = value
        else:
            retry = messagebox.askretrycancel(
                "Ошибка",
                f"Не удалось получить курс для {crypto}. Попробовать еще раз?",
            )
            if not retry:
                manual_price = simpledialog.askfloat(
                    "Ввод вручную", f"Введите курс для {crypto} в рублях:"
                )
                if manual_price is not None:
                    value = amount * manual_price
                    total_value_rub += value
                    crypto_values[crypto] = value
                else:
                    return
            else:
                all_prices_fetched = False
                break

    if all_prices_fetched:
        crypto_values["Total Value (RUB)"] = total_value_rub

        excel_file = "crypto_portfolio.xlsx"
        selected_sheet = sheet_combobox.get()

        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            if selected_sheet not in wb.sheetnames:
                ws = wb.create_sheet(selected_sheet)
            else:
                ws = wb[selected_sheet]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = selected_sheet

        df_new = pd.DataFrame([crypto_values])

        # Запись заголовка только если лист пуст
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            for col_num, column_title in enumerate(df_new.columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)

        for row in dataframe_to_rows(df_new, index=False, header=False):
            ws.append(row)

        wb.save(excel_file)
        print(f"Общая стоимость вашего портфеля: {total_value_rub:.2f} RUB")
        messagebox.showinfo("Успех", "Данные записаны в таблицу Excel!")


# Создание графического интерфейса
app = tk.Tk()
app.title("Crypto Portfolio Manager")
app.geometry("400x680")  # Устанавливаем размер окна

# Настройка шрифтов и цветов
header_font = ("Helvetica", 16, "bold")
label_font = ("Helvetica", 12)
entry_font = ("Helvetica", 12)
button_font = ("Helvetica", 12, "bold")

# Заголовок приложения
tk.Label(app, text="Crypto Portfolio Manager", font=header_font).pack(pady=10)

# Рамка для ввода данных
input_frame = tk.Frame(app)
input_frame.pack(pady=10)

tk.Label(
    input_frame, text="Введите количество криптовалют:", font=label_font
).grid(row=0, column=0, columnspan=2, pady=5)

entry_vars = {}
row = 1
for crypto in cryptos:
    tk.Label(input_frame, text=crypto, font=label_font).grid(
        row=row, column=0, padx=5, pady=5, sticky="e"
    )
    entry_vars[crypto] = tk.StringVar(value=str(portfolio[crypto]))
    tk.Entry(
        input_frame, textvariable=entry_vars[crypto], font=entry_font
    ).grid(row=row, column=1, padx=5, pady=5)
    row += 1

# Кнопки управления
button_frame = tk.Frame(app)
button_frame.pack(pady=10)

tk.Button(
    button_frame,
    text="Обновить портфель",
    command=update_portfolio,
    font=button_font,
    bg="#4CAF50",
    fg="white",
).pack(pady=5)

tk.Label(button_frame, text="Выберите лист Excel:", font=label_font).pack(
    pady=5
)
sheet_combobox = ttk.Combobox(button_frame, font=entry_font)
sheet_combobox.pack(pady=5)
update_sheet_list()

tk.Button(
    button_frame,
    text="Создать новый лист",
    command=create_new_sheet,
    font=button_font,
    bg="#FF9800",
    fg="white",
).pack(pady=5)
tk.Button(
    button_frame,
    text="Сделать запись",
    command=record_data,
    font=button_font,
    bg="#2196F3",
    fg="white",
).pack(pady=5)
tk.Button(
    button_frame,
    text="Выйти",
    command=app.quit,
    font=button_font,
    bg="#f44336",
    fg="white",
).pack(pady=5)

app.mainloop()
